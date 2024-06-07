# -------- Documentation ----------
# python +3.11.0 install
# pip install selenium
# setup chrome.exe to system variable path in Development
# Download and place to C:\ driver,  https://googlechromelabs.github.io/chrome-for-testing/#stable 

from selenium import webdriver
from selenium.common.exceptions import *
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.common.actions.action_builder import ActionBuilder
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support import ui
from time import sleep
from threading import Thread
from openpyxl import Workbook
import json
import random

def Find_Element(driver : webdriver.Chrome, by, value : str) -> WebElement:
    while True:
        try:
            element = driver.find_element(by, value)
            break
        except:
            pass
        sleep(0.1)
    return element

def Find_Elements(driver : webdriver.Chrome, by, value : str) -> list[WebElement]:
    while True:
        try:
            elements = driver.find_elements(by, value)
            if len(elements) > 0:
                break
        except:
            pass
        sleep(0.1)
    return elements

def Send_Keys(element : WebElement, content : str):
    element.clear()
    for i in content:
        element.send_keys(i)
        sleep(0.1)

service = Service(executable_path="C:\chromedriver-win64\chromedriver.exe")   
options = Options()
options.add_experimental_option("debuggerAddress", "127.0.0.1:9000")
driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=options)

wb = Workbook()
sheet = wb.active
item = ['Product Link', 'Title', 'Price', 'Processor', 'Graphics adapter', 'Memory', 'Display', 'Mainboard', 'Storage', 'Soundcard', 'Connections', 'Networking', 'Size', 'Battery', 'Operating System', 'Camera', 'Additional features', 'Weight', 'Average Score']
for i in range(1, 20):
    sheet.cell(row = 1, column = i).value = item[i-1]

start_row = 2
specification_col = 4
item_num = 0

# get links 1
# driver.get('https://www.notebookcheck.net/Rating-display.2679.0.html')
# product_items = driver.find_element(By.CLASS_NAME, "tx-nbc2fe-pi1").find_elements(By.TAG_NAME, 'p')
# output = []

# for product_item in product_items[1:]:
#     product_url = product_item.find_element(By.TAG_NAME, 'a').get_attribute('href')
#     output.append({"product link" : product_url})
    
# with open('item_url.json', 'w') as file:
#     json.dump(output, file)

# get links 2
driver.get('https://www.notebookcheck.net/Comprehensive-list-of-all-notebooks-featuring-displays-with-high-refresh-rates.522463.0.html')
product_items = driver.find_element(By.CLASS_NAME, "dataTable-container").find_elements(By.TAG_NAME, 'tr')
output = []
print(len(product_items))
for product_item in product_items:
    try:
        product_url = product_item.find_element(By.TAG_NAME, 'a').get_attribute('href')
        output.append({"product link" : product_url})
    except:
        pass
    
with open('item_url.json', 'w') as file:
    json.dump(output, file)

# scrape data from item_url.json
# with open('item_url_diff.json', 'r') as file:
#     items = json.load(file)
    
# for item_index, item in enumerate(items):
#     driver.get(item["product link"])
#     try:
#         title = driver.find_element(By.CLASS_NAME, 'tx-nbc2fe-pi1').find_element(By.TAG_NAME, 'h1').text
        
#         #link and title
#         item_num += 1
#         sheet.cell(row = start_row, column = 1).value = item["product link"]
#         print('-' * 20, item_num, '-' * 20)
#         print(f'Product URL : {item["product link"]}')
#         sheet.cell(row = start_row, column = 2).value = title
#         print(f'Title : {title}')
        
#         specifications = driver.find_element(By.CLASS_NAME, 'specs_whole ').find_elements(By.CLASS_NAME, 'specs_element')
#         for specification in specifications:
#             if "Price" in specification.text:
#                 price = specification.text.replace('Price', '').replace('\n', '').replace(' ,', ',').strip()
#                 sheet.cell(row = start_row, column = 3).value = price
#                 print(f'Price : {price}')
                
#             elif "Processor" in specification.text:
#                 processor = 'Processor : ' + specification.text.replace('Processor', '').replace('\n', '').replace(' ,', ',').strip()
#                 sheet.cell(row = start_row, column = specification_col).value = processor
#                 print(processor)
                
#             elif "Graphics adapter" in specification.text:
#                 graphics_adapter = 'Graphics adapter : ' + specification.text.replace('Graphics adapter', '').replace('\n', '').replace(' ,', ',').strip()
#                 sheet.cell(row = start_row, column = specification_col + 1).value = graphics_adapter
#                 print(graphics_adapter)
                
#             elif "Memory" in specification.text:
#                 memory = 'Memory : ' + specification.text.replace('Memory', '').replace('\n', '').replace(' ,', ',').strip()
#                 sheet.cell(row = start_row, column = specification_col + 2).value = memory
#                 print(memory)
                
#             elif "Display" in specification.text:
#                 display = 'Display : ' + specification.text.replace('Display', '').replace('\n', '').replace(' ,', ',').strip()
#                 sheet.cell(row = start_row, column = specification_col + 3).value = display
#                 print(display)
                
#             elif "Mainboard" in specification.text:
#                 mainboard = 'Mainboard : ' + specification.text.replace('Mainboard', '').replace('\n', '').replace(' ,', ',').strip()
#                 sheet.cell(row = start_row, column = specification_col + 4).value = mainboard
#                 print(mainboard)
                
#             elif "Storage" in specification.text:
#                 storage = 'Storage : ' + specification.text.replace('Storage', '').replace('\n', '').replace(' ,', ',').strip()
#                 sheet.cell(row = start_row, column = specification_col + 5).value = storage
#                 print(storage)
                
#             elif "Soundcard" in specification.text:
#                 soundcard = 'Soundcard : ' + specification.text.replace('Soundcard', '').replace('\n', '').replace(' ,', ',').strip()
#                 sheet.cell(row = start_row, column = specification_col + 6).value = soundcard
#                 print(soundcard)
                
#             elif "Connections" in specification.text:
#                 connections = 'Connections : ' + specification.text.replace('Connections', '').replace('\n', '').replace(' ,', ',').strip()
#                 sheet.cell(row = start_row, column = specification_col + 7).value = connections
#                 print(connections)
                
#             elif "Networking" in specification.text:
#                 networking = 'Networking : ' + specification.text.replace('Networking', '').replace('\n', '').replace(' ,', ',').strip()
#                 sheet.cell(row = start_row, column = specification_col + 8).value = networking
#                 print(networking)
                
#             elif "Size" in specification.text:
#                 size = 'Size : ' + specification.text.replace('Size', '').replace('\n', '').replace(' ,', ',').strip()
#                 sheet.cell(row = start_row, column = specification_col + 9).value = size
#                 print(size)
                
#             elif "Battery" in specification.text:
#                 battery = 'Battery : ' + specification.text.replace('Battery', '').replace('\n', '').replace(' ,', ',').strip()
#                 sheet.cell(row = start_row, column = specification_col + 10).value = battery
#                 print(battery)
                
#             elif "Operating System" in specification.text:
#                 operating_system = 'Operating System : ' + specification.text.replace('Operating System', '').replace('\n', '').replace(' ,', ',').strip()
#                 sheet.cell(row = start_row, column = specification_col + 11).value = operating_system
#                 print(operating_system)
                
#             elif "Camera" in specification.text:
#                 camera = 'Camera : ' + specification.text.replace('Camera', '').replace('\n', '').replace(' ,', ',').strip()
#                 sheet.cell(row = start_row, column = specification_col + 12).value = camera
#                 print(camera)
                
#             elif "Additional features" in specification.text:
#                 additional_features = 'Additional features : ' + specification.text.replace('Additional features', '').replace('\n', '').replace(' ,', ',').strip()
#                 sheet.cell(row = start_row, column = specification_col + 13).value = additional_features
#                 print(additional_features)
                
#             elif "Weight" in specification.text:
#                 weight = 'Weight : ' + specification.text.replace('Weight', '').replace('\n', '').replace(' ,', ',').strip()
#                 sheet.cell(row = start_row, column = specification_col + 14).value = weight
#                 print(weight)
        
#         # Average Score
#         try:
#             average_score = driver.find_element(By.CLASS_NAME, 'hreview-aggregate').text.split('\n')[0]
#             sheet.cell(row = start_row, column = specification_col + 15).value = average_score
#             print(average_score)
#         except:
#             pass        
#         wb.save(f'output_rating.xlsx')
#         start_row += 1
        
#     except:
#         pass