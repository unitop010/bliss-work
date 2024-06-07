# -------- Documentation ----------
# python +3.11.0 install
# pip install selenium
# setup chrome.exe to system variable path in Development
# Download and place to C:\ driver,  https://googlechromelabs.github.io/chrome-for-testing/#stable 

from selenium import webdriver
from selenium.common.exceptions import *
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.common.actions.action_builder import ActionBuilder
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support import ui
from time import sleep
from threading import Thread
from openpyxl import Workbook
import json
import random

def Find_Element(driver : webdriver.Chrome, by, value : str) -> WebElement:
    while True:
        try:
            element = driver.find_element(by, value)
            break
        except:
            pass
        sleep(0.1)
    return element

def Find_Elements(driver : webdriver.Chrome, by, value : str) -> list[WebElement]:
    while True:
        try:
            elements = driver.find_elements(by, value)
            if len(elements) > 0:
                break
        except:
            pass
        sleep(0.1)
    return elements

def Send_Keys(element : WebElement, content : str):
    element.clear()
    for i in content:
        element.send_keys(i)
        sleep(0.1)

service = Service(executable_path="C:\chromedriver-win64\chromedriver.exe")   
options = Options()
options.add_experimental_option("debuggerAddress", "127.0.0.1:9000")
driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=options)

wb = Workbook()
sheet = wb.active
item = ['Product Link', 'Title', 'Price', 'Part Number', 'Compatible Laptop', 'Compatible Part Number', 'Size', 'Resolution', 'Touch','Connector', 'Brightness', 'Refresh Rate', 'Backlight', 'Aspect Ratio', 'Contrast Ratio', 'Color Gamut', 'Surface',  'Condition', 'Dead Pixel Policy', 'Warranty']
for i in range(1, 21):
    sheet.cell(row = 1, column = i).value = item[i-1]

start_row = 2
for id in range(1, 417):
    driver.get(f'https://www.sourceitstore.ca/product-category/acer-parts/page/{id}/?per_page=24')
    product_items = driver.find_elements(By.CLASS_NAME, "product-wrapper")
    output = []
    print(product_items)
    for product_item in product_items:
        product_url = product_item.find_element(By.TAG_NAME, 'a').get_attribute('href')
        print(product_url)

    for item_index, item in enumerate(output):
        driver.get(item["product link"])
        try:
            short_description = driver.find_element(By.CLASS_NAME, 'woocommerce-product-details__short-description').find_element(By.TAG_NAME, 'p').text
            short_description_lower = short_description.lower()
            
            #link and short-description 
            if "display" in short_description_lower or "lcd" in short_description_lower or "led" in short_description_lower or "panel" in short_description_lower or "monitor" in short_description_lower: 
                # sheet.cell(row = start_row, column = 1).value = item["product link"]
                # sheet.cell(row = start_row, column = 3).value = short_description
                # print(f'short_description : {short_description}')
                output.append({"product link" : product_url})
                
                # title
                # try:
                #     title = driver.find_element(By.CLASS_NAME, 'product_title').text
                #     sheet.cell(row = start_row, column = 2).value = title
                #     print(f'title : {title}')
                # except:
                #     pass
                
                # price
                # try:
                #     price = driver.find_element(By.CLASS_NAME, 'summary-inner').find_element(By.CLASS_NAME, 'woocommerce-Price-amount').text
                #     sheet.cell(row = start_row, column = 4).value = price
                #     print(f'price : {price}')
                # except:
                #     pass
                
                # #other data
                # try:
                #     # contents = driver.find_element(By.CLASS_NAME, 'woocommerce-Tabs-panel--description').text.split('\n')
                #     contents = driver.find_element(By.CLASS_NAME, 'woocommerce-Tabs-panel--description').find_elements(By.TAG_NAME, 'p')
                #     specification_column = 7
                #     for content in contents:
                #         if "feature" in content.text.lower() or "this is a" in content.text.lower() or "identify your" in content.text.lower() or "if you do" in content.text.lower() or "how to check" in content.text.lower() or "before placing" in content.text.lower() or "please select" in content.text.lower() or "it is" in content.text.lower() or "all four part" in content.text.lower() or "in the order" in content.text.lower():
                #             continue
                #         elif "1. " not in content.text.lower():
                #             if "part number" in content.text.lower():
                #                 if "compatible" in content.text.lower():
                #                     sheet.cell(row = start_row, column = 6).value = content.text
                #                     print(f'compatible part number : {content.text}')
                #                 else :
                #                     sheet.cell(row = start_row, column = 4).value = content.text
                #                     print(f'part number : {content.text}')
                                    
                #             elif "compatible laptop" in content.text.lower():
                #                 sheet.cell(row = start_row, column = 5).value = content.text
                #                 print(f'compatible laptop : {content.text}')
                                
                #             elif "specification" in content.text.lower():
                #                 specifications = content.text.split('\n')
                #                 for specification in specifications:
                #                     sheet.cell(row = start_row, column = specification_column).value = specification
                #                     print(f'specifications{specification_column} : {specification}')         
                #                     specification_column += 1                       
                #         else :
                #             break
                            
                # except:
                #     pass
                
                # wb.save('output.xlsx')
                # start_row += 1
                
                with open('item_url.json', 'w') as file:
                    json.dump(output, file)
        except:
            pass
        sleep(random.uniform(0.1, 1))