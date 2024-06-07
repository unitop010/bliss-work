# -------- Documentation ----------
# python +3.11.0 install
# pip install selenium
# setup chrome.exe to system variable path in Development
# Download and place to C:\ driver,  https://googlechromelabs.github.io/chrome-for-testing/#stable 

from selenium import webdriver
from selenium.common.exceptions import *
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.common.actions.action_builder import ActionBuilder
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support import ui
from time import sleep
from threading import Thread
from openpyxl import Workbook
import openpyxl
from openpyxl.utils import get_column_letter
import json

def Find_Element(driver : webdriver.Chrome, by, value : str) -> WebElement:
    while True:
        try:
            element = driver.find_element(by, value)
            break
        except:
            pass
        sleep(0.1)
    return element

def Find_Elements(driver : webdriver.Chrome, by, value : str) -> list[WebElement]:
    while True:
        try:
            elements = driver.find_elements(by, value)
            if len(elements) > 0:
                break
        except:
            pass
        sleep(0.1)
    return elements

def Send_Keys(element : WebElement, content : str):
    element.clear()
    for i in content:
        element.send_keys(i)
        sleep(0.1)

service = Service(executable_path="C:\chromedriver-win64\chromedriver.exe")   
options = Options()
options.add_experimental_option("debuggerAddress", "127.0.0.1:9000")
driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=options)

wb = openpyxl.load_workbook('output.xlsx')
sheet = wb['new']
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

target_column = 'I'
item_num = 0

for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=9, max_col=sheet.max_column):
# for row in sheet.iter_rows():
    for cell in row:
        # value = cell.value
        # print(value)
        if cell.value == None:
            item_num += 1
            print('*' * 20, item_num,'/',row, '*' * 20)
            driver.get(sheet['A' + str(cell.row)].value)
            com_models = driver.find_elements(By.CLASS_NAME, "col-lg-6")
            while cell.value == None:
                # new_sheet[target_column + str(cell.row)].value = ",".join(com_models[1:].text)
                cell.value = ",".join([x.text for x in com_models[1:]])
                print(cell.value)
            wb.save('output.xlsx')