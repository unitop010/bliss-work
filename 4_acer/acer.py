# -------- Documentation ----------
# python +3.11.0 install
# pip install selenium
# setup chrome.exe to system variable path in Development
# Download and place to C:\ driver,  https://googlechromelabs.github.io/chrome-for-testing/#stable 

from selenium import webdriver
from selenium.common.exceptions import *
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.common.actions.action_builder import ActionBuilder
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support import ui
from time import sleep
from threading import Thread
from openpyxl import Workbook
import json
import random

def Find_Element(driver : webdriver.Chrome, by, value : str) -> WebElement:
    while True:
        try:
            element = driver.find_element(by, value)
            break
        except:
            pass
        sleep(0.1)
    return element

def Find_Elements(driver : webdriver.Chrome, by, value : str) -> list[WebElement]:
    while True:
        try:
            elements = driver.find_elements(by, value)
            if len(elements) > 0:
                break
        except:
            pass
        sleep(0.1)
    return elements

def Send_Keys(element : WebElement, content : str):
    element.clear()
    for i in content:
        element.send_keys(i)
        sleep(0.1)

service = Service(executable_path="C:\chromedriver-win64\chromedriver.exe")   
options = Options()
options.add_experimental_option("debuggerAddress", "127.0.0.1:9000")
driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=options)

wb = Workbook()
sheet = wb.active
item = ['Product Link', 'Title', 'Price', 'Description', 'SKU', 'Categories', 'Brand', 'Compatible Part Numbers', 'Product Families', 'Compatible Models', 'Weight', 'Dimensions', 'Condition',  'Model']
for i in range(1, 15):
    sheet.cell(row = 1, column = i).value = item[i-1]

start_row = 2
item_num = 0
# the last number = last page + 1
for id in range(417, 418):
    driver.get(f'https://www.acerparts.ca/shop/page/{id}/?post_type=product&per_page=24')
    product_items = driver.find_elements(By.CLASS_NAME, "product-wrapper")
    output = []
    print('*' * 20, f'page {id}', '*' * 20)
    # print(product_items)
    for product_item in product_items:
        product_url = product_item.find_element(By.TAG_NAME, 'a').get_attribute('href')
        # print(product_url)
        output.append({"product link" : product_url})
        
    for item_index, item in enumerate(output):
        driver.get(item["product link"])
        try:
            short_description = driver.find_element(By.CLASS_NAME, 'woocommerce-product-details__short-description').find_element(By.TAG_NAME, 'p').text
            short_description_lower = short_description.lower()
            
            #link and short-description 
            if "display" in short_description_lower or "lcd" in short_description_lower or "led" in short_description_lower or "panel" in short_description_lower or "monitor" in short_description_lower: 
                item_num += 1
                sheet.cell(row = start_row, column = 1).value = item["product link"]
                print('-' * 20, item_num, '-' * 20)
                print(f'product URL : {item["product link"]}')
                sheet.cell(row = start_row, column = 4).value = short_description
                print(f'short_description : {short_description}')
                
                # title
                try:
                    title = driver.find_element(By.CLASS_NAME, 'product_title').text
                    sheet.cell(row = start_row, column = 2).value = title
                    print(f'title : {title}')
                except:
                    pass
                
                # price
                try:
                    price = driver.find_element(By.CLASS_NAME, 'summary-inner').find_element(By.CLASS_NAME, 'woocommerce-Price-amount').text
                    sheet.cell(row = start_row, column = 3).value = price
                    print(f'price : {price}')
                except:
                    pass
                
                # SKU
                try:
                    sku = driver.find_element(By.CLASS_NAME, 'sku_wrapper').text.split(':')
                    sheet.cell(row = start_row, column = 5).value = sku[1]
                    print(f'SKU : {sku[1]}')
                except:
                    pass
                
                # Categories
                try:
                    categories = driver.find_element(By.CLASS_NAME, 'posted_in').text.split(':')
                    sheet.cell(row = start_row, column = 6).value = categories[1]
                    print(f'Categories : {categories[1]}')
                except:
                    pass
                
                # brand
                try:
                    brand = driver.find_element(By.CLASS_NAME, 'cr_brand').text.split(':')
                    sheet.cell(row = start_row, column = 7).value = brand[1]
                    print(f'Brand : {brand[1]}')
                except:
                    pass
                
                # other data
                try:
                    other_datum = driver.find_elements(By.CLASS_NAME, 'wc-tab-inner')
                    # description
                    descriptions = other_datum[0].find_elements(By.TAG_NAME, 'p')
                    for description in descriptions:
                        if "Compatible Part Numbers" in description.text:
                            compatible_part_numbers = description.text.split(':')
                            sheet.cell(row = start_row, column = 8).value = compatible_part_numbers[1]
                            print(f'Compatible Part Numbers : {compatible_part_numbers[1]}')
                            
                        elif "Product Families" in description.text:
                            product_families = description.text.split(':')
                            sheet.cell(row = start_row, column = 9).value = product_families[1]
                            print(f'Product Families : {product_families[1]}')
                            
                        elif "Compatible Models" in description.text:
                            compatible_models = description.text.split(':')
                            sheet.cell(row = start_row, column = 10).value = compatible_models[1]
                            print(f'Compatible Models : {compatible_models[1]}')
                    
                    # additional information
                    menus = driver.find_element(By.CLASS_NAME, 'wd-nav-tabs').find_elements(By.TAG_NAME, 'li')
                    for menu in menus:
                        if "additional information" in menu.text.lower():
                            driver.execute_script('arguments[0].click()', menu.find_element(By.TAG_NAME, 'span'))
                            add_infos = other_datum[1].find_elements(By.TAG_NAME, 'td')
                            add_infos_column = 11
                            for add_info in add_infos:
                                sheet.cell(row = start_row, column = add_infos_column).value = add_info.text
                                print(f'Additional Information[{add_infos_column}] : {add_info.text}')
                                add_infos_column += 1
                            break
                except:
                    pass
                
                wb.save('output.xlsx')
                start_row += 1
                
                with open('item_url.json', 'w') as file:
                    json.dump(output, file)
        except:
            pass
        # sleep(random.uniform(1, 1.5))