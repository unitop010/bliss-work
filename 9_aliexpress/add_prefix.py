import openpyxl

# Load the Excel file
wb = openpyxl.load_workbook('aliexpress.xlsx')
sheet = wb.active

# Define the prefixes for each column
prefixes = {
    4: "Type",
    5: "Material",
    6: "Screen",
    7: "Model Number",
    8: "LCD Kit",
    9: "Touch",
    10: "Display Resolution",
    11: "Brand",
    12: "Origin",
    13: "Certification",
    14: "Color",
    15: "Resolution",
    16: "Size",
    17: "Compatible Brand",
    18: "Warranty",
    19: "QC",
    20: "Item Type"
}

# Iterate through each column and add the corresponding prefix to non-blank cells
for col_num, prefix in prefixes.items():
    for cell in sheet.iter_cols(min_row=2, max_row=sheet.max_row, min_col=col_num, max_col=col_num):
        for cell_in_col in cell:
            if cell_in_col.value is not None:
                cell_in_col.value = f"{prefix} : {cell_in_col.value}"
# Save the modified Excel file
print('Success')
wb.save('modified_excel_file.xlsx')