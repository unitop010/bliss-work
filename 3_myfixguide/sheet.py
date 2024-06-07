import openpyxl
from openpyxl.utils import get_column_letter

# Load the Excel file
wb = openpyxl.load_workbook('output_basic.xlsx')

# Access the sheet
sheet = wb['Sheet']
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active
# Iterate through the rows and move cells with specific text to the target column
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=3, max_col=sheet.max_column):
# for row in sheet.iter_rows():
  for cell in row:
    # value = cell.value
    # print(value)
    
    if cell.value is not None and 'compatible laptop' in cell.value.lower():
      target_column = 'E'
      new_sheet[target_column + str(cell.row)].value = cell.value.replace('Compatible', '').replace('compatible', '').replace('Laptops', '').replace('laptops', '').replace('Laptop', '').replace('laptop', '').replace(':', '').strip()
        
    elif cell.value is not None and 'compatible part number' in cell.value.lower():
      target_column = 'F'
      new_sheet[target_column + str(cell.row)].value = cell.value.replace('Compatible', '').replace('compatible', '').replace('Part', '').replace('Numbers', '').replace('Number', '').replace(':', '').strip()
        
    elif cell.value is not None and 'part number' in cell.value.lower():
      target_column = 'D'
      new_sheet[target_column + str(cell.row)].value = cell.value.replace('Complete', '').replace('Sub', '').replace('Part', '').replace('Numbers', '').replace('Number', '').replace(':', '').strip()
        
    elif cell.value is not None and 'size' in cell.value.lower():
      target_column = 'G'
      new_sheet[target_column + str(cell.row)].value = cell.value.replace('Size', '').replace('size', '').replace(':', '').strip()
        
    elif cell.value is not None and 'resolution' in cell.value.lower():
      target_column = 'H'
      new_sheet[target_column + str(cell.row)].value = cell.value.replace('Resolution', '').replace('resolution', '').replace(':', '').strip()
        
    elif cell.value is not None and 'touch' in cell.value.lower():
      target_column = 'I'
      new_sheet[target_column + str(cell.row)].value = cell.value.replace('Touch:', '').replace('Touch-Function:', '').replace(':', '').strip()
        
    elif cell.value is not None and 'connector' in cell.value.lower():
      target_column = 'J'
      new_sheet[target_column + str(cell.row)].value = cell.value.replace('Connector', '').replace('Video', '').replace(':', '').strip()
        
    elif cell.value is not None and 'brightness' in cell.value.lower():
      target_column = 'K'
      new_sheet[target_column + str(cell.row)].value = cell.value.replace('Brightness', '').replace('brightness', '').replace(':', '').strip()
        
    elif cell.value is not None and 'refresh rate' in cell.value.lower():
      target_column = 'L'
      new_sheet[target_column + str(cell.row)].value = cell.value.replace('Refresh', '').replace('Rate', '').replace(':', '').strip()
        
    elif cell.value is not None and 'backlight' in cell.value.lower():
      target_column = 'M'
      new_sheet[target_column + str(cell.row)].value = cell.value.replace('Backlight', '').replace('Type', '').replace('type', '').replace(':', '').strip()
        
    elif cell.value is not None and 'aspect aspect' in cell.value.lower():
      target_column = 'N'
      new_sheet[target_column + str(cell.row)].value = cell.value.replace('Aspect', '').replace('Ratio', '').replace(':', '').strip()
      
    elif cell.value is not None and 'contrast aspect' in cell.value.lower():
      target_column = 'O'
      new_sheet[target_column + str(cell.row)].value = cell.value.replace('Contrast', '').replace('Ratio', '').replace(':', '').strip()
      
    elif cell.value is not None and 'color gamut' in cell.value.lower():
      target_column = 'P'
      new_sheet[target_column + str(cell.row)].value = cell.value.replace('Color', '').replace('Gamut', '').replace(':', '').strip()
              
    elif cell.value is not None and 'surface' in cell.value.lower():
      target_column = 'Q'
      new_sheet[target_column + str(cell.row)].value = cell.value.replace('Surface', '').replace('Type', '').replace('type', '').replace(':', '').strip()
            
    elif cell.value is not None and 'condition' in cell.value.lower():
      target_column = 'R'
      new_sheet[target_column + str(cell.row)].value = cell.value.replace('Condition', '').replace('condition', '').replace(':', '').strip()
            
    elif cell.value is not None and 'dead pixel policy' in cell.value.lower():
      target_column = 'S'
      new_sheet[target_column + str(cell.row)].value = cell.value.replace('Dead Pixel Policy', '').replace(':', '').strip()      
            
    elif cell.value is not None and 'warranty' in cell.value.lower():
      target_column = 'T'
      new_sheet[target_column + str(cell.row)].value = cell.value.replace('Warranty', '').replace(':', '').strip()      
      
print('Success')
# Save the modified workbook
new_wb.save('modified_file.xlsx')
# wb.save('modified_file.xlsx')

