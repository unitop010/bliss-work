import openpyxl
from openpyxl.utils import get_column_letter

# Load the Excel file
wb = openpyxl.load_workbook('output04-1.xlsx')

# Access the sheet
sheet = wb['Sheet']
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active
# Iterate through the rows and move cells with specific text to the target column
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=7, max_col=sheet.max_column):
# for row in sheet.iter_rows():
  for cell in row:
    # value = cell.value
    # print(value)
    
    if cell.value is not None and 'condition' in cell.value.lower():
      target_column = 'G'
      new_sheet[target_column + str(cell.row)].value = cell.value.replace('Condition', '').replace('condition', '').replace('#', '').replace(':', '').strip()
      cell.value = None
        
    elif cell.value is not None and 'brand' in cell.value.lower():
      target_column = 'H'
      new_sheet[target_column + str(cell.row)].value = cell.value.replace('Brand', '').replace('brand', '').replace('BRAND', '').replace('#', '').replace(':', '').strip()
      cell.value = None
        
    elif cell.value is not None and 'size' in cell.value.lower():
      target_column = 'M'
      new_sheet[target_column + str(cell.row)].value = cell.value.replace('Size', '').replace('size', '').replace('SIZE', '').replace('#', '').replace(':', '').strip()
      cell.value = None
        
    elif cell.value is not None and 'resolution' in cell.value.lower():
      target_column = 'N'
      new_sheet[target_column + str(cell.row)].value = cell.value.replace('Resolution', '').replace('resolution', '').replace('Resolution', '').replace('Maximum', '').replace('#', '').replace(':', '').strip()
      cell.value = None
        
    elif cell.value is not None and 'backlight' in cell.value.lower():
      target_column = 'O'
      new_sheet[target_column + str(cell.row)].value = cell.value.replace('Backlight', '').replace('backlight', '').replace('Type', '').replace('type', '').replace('LCD', '').replace('Lamp', '').replace('Illumination', '').replace('#', '').replace(':', '').strip()
      cell.value = None
                    
    elif cell.value is not None and 'surface' in cell.value.lower():
      target_column = 'P'
      new_sheet[target_column + str(cell.row)].value = cell.value.replace('Surface', '').replace('surface', '').replace('#', '').replace(':', '').strip()
      cell.value = None
                    
    elif cell.value is not None and 'lamp' in cell.value.lower():
      target_column = 'Q'
      new_sheet[target_column + str(cell.row)].value = cell.value.replace('Lamp', '').replace('Screen', '').replace('Type', '').replace('type', '').replace('#', '').replace('_', '').replace(':', '').strip()
      cell.value = None
        
    elif cell.value is not None and 'data cable pins' in cell.value.lower():
      target_column = 'R'
      new_sheet[target_column + str(cell.row)].value = cell.value.replace('Data Cable Pins', '').replace('#', '').replace('_', '').replace(':', '').strip()
      cell.value = None
        
    elif cell.value is not None and 'connector' in cell.value.lower():
      target_column = 'R'
      if cell.value == 'Connector' or cell.value == 'Connector:':
        next_column_index = cell.column + 1
        next_column_letter = get_column_letter(next_column_index)
        next_cell = next_column_letter + str(cell.row)
        new_sheet[target_column + str(cell.row)].value = next_cell.replace('Connector', '').replace('#', '').replace('_', '').replace(':', '').strip()
      else:
        new_sheet[target_column + str(cell.row)].value = cell.value.replace('Connector', '').replace('#', '').replace('_', '').replace(':', '').strip()
      cell.value = None
    
    elif cell.value is not None and 'aspect ratio' in cell.value.lower():
      target_column = 'S'
      new_sheet[target_column + str(cell.row)].value = cell.value.replace('Aspect', '').replace('aspect', '').replace('Ratio', '').replace('ratio', '').replace('#', '').replace('_', '').replace(':', '').strip()
      cell.value = None
      
    elif cell.value is not None and 'packing' in cell.value.lower():
      target_column = 'T'
      new_sheet[target_column + str(cell.row)].value = cell.value.replace('Packing', '').replace('packing', '').replace('#', '').replace('_', '').replace(':', '').strip()
      cell.value = None
      
    elif cell.value is not None and 'description' in cell.value.lower():
      target_column = 'U'
      if cell.value.lower() == 'description' or cell.value.lower() == 'description:':
        next_column_index = cell.column + 1
        next_column_letter = get_column_letter(next_column_index)
        next_cell = next_column_letter + str(cell.row)
        new_sheet[target_column + str(cell.row)].value = next_cell.replace('Description', '').replace('description', '').replace('#', '').replace('_', '').replace(':', '').strip()
      else:
        new_sheet[target_column + str(cell.row)].value = cell.value.replace('Description', '').replace('description', '').replace('#', '').replace('_', '').replace(':', '').strip()
      cell.value = None
    
    elif cell.value is not None and 'warranty' in cell.value.lower():
      target_column = 'V'
      if cell.value.lower() == 'warranty' or cell.value.lower() == 'warranty :':
        next_column_index = cell.column + 1
        next_column_letter = get_column_letter(next_column_index)
        next_cell = next_column_letter + str(cell.row)
        new_sheet[target_column + str(cell.row)].value = next_cell.replace('Warranty', '').replace('warranty', '').replace('#', '').replace('_', '').replace(':', '').strip()
      else:
        new_sheet[target_column + str(cell.row)].value = cell.value.replace('Warranty', '').replace('warranty', '').replace('#', '').replace('_', '').replace(':', '').strip()
      cell.value = None
      
    elif cell.value is not None and 'fru' in cell.value.lower():
      target_column = 'J'
      new_sheet[target_column + str(cell.row)].value = cell.value.replace('Fru', '').replace('FRU', '').replace('Number', '').replace('number', '').replace('#', '').replace('_', '').replace(':', '').strip()
      cell.value = None
    
    elif cell.value is not None and 'mpn' in cell.value.lower():
      target_column = 'K'
      new_sheet[target_column + str(cell.row)].value = cell.value.replace('MPN', '').replace('#', '').replace('_', '').replace(':', '').strip()
      cell.value = None
    
    elif cell.value is not None and 'model' in cell.value.lower():
      target_column = 'K'
      new_sheet[target_column + str(cell.row)].value = cell.value.replace('Model', '').replace('Fit', '').replace('Number', '').replace('number', '').replace('No', '').replace('#', '').replace('_', '').replace(':', '').replace('.', '').strip()
      cell.value = None
    
    elif cell.value is not None and 'part' in cell.value.lower():
      target_column = 'L'
      new_sheet[target_column + str(cell.row)].value = cell.value.replace('Part', '').replace('part', '').replace('Number', '').replace('number', '').replace('#', '').replace('_', '').replace(':', '').strip()
      cell.value = None
    
    elif cell.value is not None and 'type' in cell.value.lower():
      target_column = 'I'
      new_sheet[target_column + str(cell.row)].value = cell.value.replace('Type', '').replace('type', '').replace('#', '').replace('_', '').replace(':', '').strip()
      cell.value = None
    
print('Success')
# Save the modified workbook
new_wb.save('modified_file.xlsx')
# wb.save('modified_file.xlsx')

