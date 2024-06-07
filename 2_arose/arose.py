# -------- Documentation ----------
# python +3.11.0 install
# pip install selenium
# setup chrome.exe to system variable path in Development
# Download and place to C:\ driver,  https://googlechromelabs.github.io/chrome-for-testing/#stable 

from selenium import webdriver
from selenium.common.exceptions import *
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.common.actions.action_builder import ActionBuilder
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support import ui
from time import sleep
from threading import Thread
from openpyxl import Workbook
import json

def Find_Element(driver : webdriver.Chrome, by, value : str) -> WebElement:
    while True:
        try:
            element = driver.find_element(by, value)
            break
        except:
            pass
        sleep(0.1)
    return element

def Find_Elements(driver : webdriver.Chrome, by, value : str) -> list[WebElement]:
    while True:
        try:
            elements = driver.find_elements(by, value)
            if len(elements) > 0:
                break
        except:
            pass
        sleep(0.1)
    return elements

def Send_Keys(element : WebElement, content : str):
    element.clear()
    for i in content:
        element.send_keys(i)
        sleep(0.1)

service = Service(executable_path="C:\chromedriver-win64\chromedriver.exe")   
options = Options()
options.add_experimental_option("debuggerAddress", "127.0.0.1:9000")
driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=options)

wb = Workbook()
sheet = wb.active
item = ['Product Link', 'Title', 'Price', 'Commodity Name', 'Commodity ID', 'Compatibility', 'Condition', 'Brand', 'Type', 'FRU', 'Model Part Number', 'Part Number', 'Size', 'Resolution(MAX)', 'Backlight Type', 'Surface', 'Lamp', 'Connector', 'Aspect Ratio', 'Packing', 'Description', 'Warranty']
for i in range(1, 23):
    sheet.cell(row = 1, column = i).value = item[i-1]

start_row = 2
for id in range(0, 23):
    driver.get(f'https://www.aroseks.com/product/p-{id * 99}-99.html')
    # product_items = Find_Element(driver, By.CLASS_NAME, 'product-listing-list').find_elements(By.CLASS_NAME, "product-item")
    product_items = driver.find_elements(By.CLASS_NAME, "cbox-1")
    output = []
    print(product_items)
    for product_item in product_items:
        product_url = product_item.find_element(By.TAG_NAME, 'a').get_attribute('href')
        print(product_url)
        output.append({"product link" : product_url})

    for item_index, item in enumerate(output):
        driver.get(item["product link"])
        try:
            title = driver.find_element(By.CLASS_NAME, 's_subtitle').text
            title_lower = title.lower()
            
            #link and title 
            if "display" in title_lower or "lcd" in title_lower or "led" in title_lower or "panel" in title_lower or "monitor" in title_lower:
                sheet.cell(row = start_row, column = 1).value = item["product link"]
                sheet.cell(row = start_row, column = 2).value = title
                print(f'title : {title}')
            
            # sheet.cell(row = start_row, column = 2).value = " ".join(title[2:])
            # print(f'title : {" ".join(title[2:])}')
            
                # price
                try:
                    price = driver.find_element(By.CLASS_NAME, 'e_text-96').text
                    if price == "":
                        price = driver.find_element(By.CLASS_NAME, 'e_text-99').text
                        price = "SKU " + price
                    sheet.cell(row = start_row, column = 3).value = price
                    print(f'price : {price}')
                except:
                    pass
                
                # commodity name & ID
                try:
                    commodity = driver.find_elements(By.CLASS_NAME, 'p_attrValue')
                    sheet.cell(row = start_row, column = 4).value = commodity[0].text
                    sheet.cell(row = start_row, column = 5).value = commodity[1].text
                    print(f'commodity_name : {commodity[0].text}   commodity_ID : {commodity[1].text}')
                except:
                    pass
                
                #other data
                try:
                    contents = driver.find_element(By.CLASS_NAME, 'p_tabContent').text.split('\n')
                    other_column = 7
                    compatibility = []
                    is_compatible = False
                    description = ["Description :"]
                    is_description_entered = False
                    for content in contents:
                        if content != "":
                            if "Key words" not in content:
                                if is_description_entered:
                                    description.append(content)
                                    is_description_entered = False
                                if not is_compatible:
                                    sheet.cell(row = start_row, column = other_column).value = content
                                    print(f'other data{other_column} : {content}')
                                    other_column += 1
                            else:
                                break
                        if content.lower() == 'description':
                            is_description_entered = True
                        if is_compatible:
                            compatibility.append(content)
                        if "compat" in content.lower() and "lity" in content.lower() or "compatible model" in content.lower():
                            is_compatible = True
                    sheet.cell(row = start_row, column = 6).value = ", ".join(compatibility).strip()
                    print(f'Compatibility: {compatibility}')
                        
                except:
                    pass
                wb.save('output.xlsx')
                start_row += 1
                
                with open('item_url.json', 'w') as file:
                    json.dump(output, file)
        except:
            pass
        