# -------- Documentation ----------
# python +3.11.0 install
# pip install selenium
# setup chrome.exe to system variable path in Development
# Download and place to C:\ driver,  https://googlechromelabs.github.io/chrome-for-testing/#stable 

from selenium import webdriver
from selenium.common.exceptions import *
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.common.actions.action_builder import ActionBuilder
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support import ui
from time import sleep
from threading import Thread
from openpyxl import Workbook
import json
import random

def Find_Element(driver : webdriver.Chrome, by, value : str) -> WebElement:
    while True:
        try:
            element = driver.find_element(by, value)
            break
        except:
            pass
        sleep(0.1)
    return element

def Find_Elements(driver : webdriver.Chrome, by, value : str) -> list[WebElement]:
    while True:
        try:
            elements = driver.find_elements(by, value)
            if len(elements) > 0:
                break
        except:
            pass
        sleep(0.1)
    return elements

def Send_Keys(element : WebElement, content : str):
    element.clear()
    for i in content:
        element.send_keys(i)
        sleep(0.1)

service = Service(executable_path="C:\chromedriver-win64\chromedriver.exe")   
options = Options()
options.add_experimental_option("debuggerAddress", "127.0.0.1:9000")
driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=options)

wb = Workbook()
sheet = wb.active
item = ['Product Link', 'Title', 'Title_2', 'Display', 'Name', 'Price']
for i in range(1, 7):
    sheet.cell(row = 1, column = i).value = item[i-1]

start_row = 2
specification_col = 4
item_num = 0
# the last number = last page + 1
for id in range(1, 101):
    # divided links
    #1//~500/5523: https://laptopmedia.com/specs/?current=n_2_n&size=n_100_n&filters%5B0%5D%5Bfield%5D=price_list&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bto%5D=n_500_n&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bfrom%5D=n_1_n&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bname%5D=Under%20%24500&filters%5B0%5D%5Btype%5D=any
    #2//500~650/6856: https://laptopmedia.com/specs/?current=n_2_n&size=n_100_n&filters%5B0%5D%5Bfield%5D=price_list&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bto%5D=n_650_n&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bfrom%5D=n_500_n&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bname%5D=%24500%20-%20%24800&filters%5B0%5D%5Btype%5D=any
    #3//650~770/6709: https://laptopmedia.com/specs/?current=n_2_n&size=n_100_n&filters%5B0%5D%5Bfield%5D=price_list&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bto%5D=n_770_n&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bfrom%5D=n_650_n&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bname%5D=%24500%20-%20%24800&filters%5B0%5D%5Btype%5D=any
    #4//770~850/5002: https://laptopmedia.com/specs/?current=n_2_n&size=n_100_n&filters%5B0%5D%5Bfield%5D=price_list&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bto%5D=n_850_n&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bfrom%5D=n_770_n&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bname%5D=%24500%20-%20%24800&filters%5B0%5D%5Btype%5D=any
    #5//850~950/5730: https://laptopmedia.com/specs/?current=n_2_n&size=n_100_n&filters%5B0%5D%5Bfield%5D=price_list&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bto%5D=n_950_n&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bfrom%5D=n_850_n&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bname%5D=%24500%20-%20%24800&filters%5B0%5D%5Btype%5D=any
    #6//950~1050/5202: https://laptopmedia.com/specs/?current=n_2_n&size=n_100_n&filters%5B0%5D%5Bfield%5D=price_list&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bto%5D=n_1050_n&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bfrom%5D=n_950_n&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bname%5D=%24500%20-%20%24800&filters%5B0%5D%5Btype%5D=any
    #7//1050~1180/5734: https://laptopmedia.com/specs/?current=n_2_n&size=n_100_n&filters%5B0%5D%5Bfield%5D=price_list&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bto%5D=n_1180_n&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bfrom%5D=n_1050_n&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bname%5D=%24500%20-%20%24800&filters%5B0%5D%5Btype%5D=any
    #8//1180~1300/6103: https://laptopmedia.com/specs/?current=n_2_n&size=n_100_n&filters%5B0%5D%5Bfield%5D=price_list&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bto%5D=n_1300_n&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bfrom%5D=n_1180_n&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bname%5D=%24500%20-%20%24800&filters%5B0%5D%5Btype%5D=any
    #9//1300~1450/6053: https://laptopmedia.com/specs/?current=n_2_n&size=n_100_n&filters%5B0%5D%5Bfield%5D=price_list&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bto%5D=n_1450_n&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bfrom%5D=n_1300_n&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bname%5D=%24500%20-%20%24800&filters%5B0%5D%5Btype%5D=any
    #10//1450~1600/5387: https://laptopmedia.com/specs/?current=n_2_n&size=n_100_n&filters%5B0%5D%5Bfield%5D=price_list&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bto%5D=n_1600_n&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bfrom%5D=n_1450_n&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bname%5D=%24500%20-%20%24800&filters%5B0%5D%5Btype%5D=any
    #11//1600~1800/5476: https://laptopmedia.com/specs/?current=n_2_n&size=n_100_n&filters%5B0%5D%5Bfield%5D=price_list&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bto%5D=n_1800_n&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bfrom%5D=n_1600_n&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bname%5D=%24500%20-%20%24800&filters%5B0%5D%5Btype%5D=any
    #12//1800~2100/5680: https://laptopmedia.com/specs/?current=n_2_n&size=n_100_n&filters%5B0%5D%5Bfield%5D=price_list&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bto%5D=n_2100_n&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bfrom%5D=n_1800_n&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bname%5D=%24500%20-%20%24800&filters%5B0%5D%5Btype%5D=any
    #13//2100~2800/6202: https://laptopmedia.com/specs/?current=n_2_n&size=n_100_n&filters%5B0%5D%5Bfield%5D=price_list&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bto%5D=n_2800_n&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bfrom%5D=n_2100_n&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bname%5D=%24500%20-%20%24800&filters%5B0%5D%5Btype%5D=any
    #14//2800~/7789: https://laptopmedia.com/specs/?current=n_2_n&size=n_100_n&filters%5B0%5D%5Bfield%5D=price_list&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bfrom%5D=n_2800_n&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bname%5D=Over%20%241500&filters%5B0%5D%5Btype%5D=any
    
    driver.get(f'https://laptopmedia.com/specs/?current=n_{id}_n&size=n_100_n&filters%5B0%5D%5Bfield%5D=price_list&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bto%5D=n_900_n&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bfrom%5D=n_750_n&filters%5B0%5D%5Bvalues%5D%5B0%5D%5Bname%5D=%24500%20-%20%24800&filters%5B0%5D%5Btype%5D=any')
    
    sleep(3)
    
    # click consent button
    try:
        driver.execute_script('arguments[0].click()', driver.find_element(By.CLASS_NAME, 'fc-cta-consent'))
    except:
        pass
    
    product_items = driver.find_element(By.CLASS_NAME, "sui-layout-main-body").find_elements(By.TAG_NAME, "li")
    
    # Wait for the element to be present
    # wait = WebDriverWait(driver, 10)
    # element = wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'sui-layout-main-body')))
    # Once the element is present, find the nested elements
    # product_items = element.find_elements(By.TAG_NAME, "li")

    output = []
    print('#' * 20, f'page {id}', '#' * 20)
    for product_item in product_items:
        product_url = product_item.find_element(By.TAG_NAME, 'a').get_attribute('href')
        output.append({"product link" : product_url})
        
    for item_index, item in enumerate(output):
        driver.execute_script("window.localStorage.clear();")
        driver.execute_script("window.sessionStorage.clear();")
        driver.execute_script("window.indexedDB.deleteDatabase('websql');")
        
        # driver.get(item["product link"])
        driver.get('https://laptopmedia.com/series/alienware-15-r3/')
        try:
            #link and title
            item_num += 1
            sheet.cell(row = start_row, column = 1).value = item["product link"]
            print('-' * 20, item_num, '-' * 20)
            print(f'Product URL : {item["product link"]}')
            try:
                title = driver.find_element(By.CLASS_NAME, 'portabilityTable').find_element(By.CLASS_NAME, 'bg-lm-lightGrayBlue').text.replace(',', '').replace('(', '').replace(')', '')
                sheet.cell(row = start_row, column = 2).value = title
                print(f'Title : {title}')
            except:
                pass
            
            # price
            try:
                price = driver.find_element(By.CLASS_NAME, 'lm-laptop-specs').find_element(By.CLASS_NAME, 'flex-row-reverse').find_elements(By.TAG_NAME, 'span')
                sheet.cell(row = start_row, column = 6).value = price[-1].text
                print(f'Price : {price[-1].text}')
            except:
                pass
            
            # display
            try:
                display = driver.find_element(By.CLASS_NAME, 'lm-display-model').find_element(By.CLASS_NAME, 'text-lm-lightBlue').text.replace(',', '').replace('(', '').replace(')', '')
                sheet.cell(row = start_row, column = 4).value = display
                print(f'Display : {display}')
            except:
                pass
            
            # name
            try:
                name = driver.find_element(By.ID, 'section-display').find_element(By.CLASS_NAME, 'display-specifications').find_element(By.CLASS_NAME, 'bg-lm-lightGrayBlue').text.replace(',', '').replace('(', '').replace(')', '')
                if 'Name' in name:
                    name = name.replace('Name', '').strip()
                    sheet.cell(row = start_row, column = 5).value = name
                    print(f'Name : {name}')
            except:
                pass
            
            # specification
            try:
                specifications = driver.find_element(By.ID, 'section-specs').find_elements(By.TAG_NAME, 'ul')
                for specification in specifications:
                    if "CPU" in specification.text:
                        cpu = 'CPU : ' + specification.text.replace('CPU', '').replace('\n', '').replace(' ,', ',').strip()
                        # sheet.cell(row = start_row, column = specification_col).value = cpu
                        print(cpu)
                        
                    elif "GPU" in specification.text:
                        gpu = 'GPU : ' + specification.text.replace('GPU', '').replace('\n', '').replace(' ,', ',').strip()
                        # sheet.cell(row = start_row, column = specification_col + 1).value = gpu
                        print(gpu)
                    
                    elif "Display" in specification.text:
                        display = 'Display : ' + specification.text.replace('Display', '').replace('\n', '').replace(' ,', ',').strip()
                        # sheet.cell(row = start_row, column = specification_col + 2).value = display
                        print(display)
                    
                    elif "HDD/SSD" in specification.text:
                        hdd_ssd = 'HDD/SSD : ' + specification.text.replace('HDD/SSD', '').replace('\n', '').replace(' ,', ',').strip()
                        # sheet.cell(row = start_row, column = specification_col + 3).value = hdd_ssd
                        print(hdd_ssd)
                    
                    elif "M.2 Slot" in specification.text:
                        m_2_slot = 'M.2 Slot : ' + specification.text.replace('M.2 Slot', '').replace('\n', '').replace(' ,', ',').strip()
                        # sheet.cell(row = start_row, column = specification_col + 4).value = m_2_slot
                        print(m_2_slot)
                    
                    elif "RAM" in specification.text:
                        ram = 'RAM : ' + specification.text.replace('RAM', '').replace('\n', '').replace(' ,', ',').strip()
                        # sheet.cell(row = start_row, column = specification_col + 5).value = ram
                        print(ram)
                    
                    elif "OS" in specification.text:
                        os = 'OS : ' + specification.text.replace('OS', '').replace('\n', '').replace(' ,', ',').strip()
                        # sheet.cell(row = start_row, column = specification_col + 6).value = os
                        print(os)
                    
                    elif "Body Material" in specification.text:
                        body_material = 'Body Material : ' + specification.text.replace('Body Material', '').replace('\n', '').replace(' ,', ',').strip()
                        # sheet.cell(row = start_row, column = specification_col + 7).value = body_material
                        print(body_material)
                    
                    elif "Dimensions" in specification.text:
                        dimensions = 'Dimensions : ' + specification.text.replace('Dimensions', '').replace('\n', '').replace(' ,', ',').strip()
                        # sheet.cell(row = start_row, column = specification_col + 8).value = dimensions
                        print(dimensions)
                    
                    elif "Weight" in specification.text:
                        weight = 'Weight : ' + specification.text.replace('Weight', '').replace('\n', '').replace(' ,', ',').strip()
                        # sheet.cell(row = start_row, column = specification_col + 9).value = weight
                        print(weight)
                        
            except:
                pass
            
            start_row += 1
        
        except:
            pass
    wb.save('output_laptopmedia-3.xlsx')
    
    driver.delete_all_cookies()

    # Execute JavaScript to clear browsing history
    # driver.execute_script("window.localStorage.clear();")
    # driver.execute_script("window.sessionStorage.clear();")
    # driver.execute_script("window.indexedDB.deleteDatabase('websql');")

    # Refresh the page to ensure the changes take effect
    driver.refresh()