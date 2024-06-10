# -------- Documentation ----------
# python +3.11.0 install
# pip install selenium
# setup chrome.exe to system variable path in Development
# Download and place to C:\ driver,  https://googlechromelabs.github.io/chrome-for-testing/#stable 

from selenium import webdriver
from selenium.common.exceptions import *
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.common.actions.action_builder import ActionBuilder
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support import ui
from time import sleep
from threading import Thread
from openpyxl import Workbook
import json
import random

def Find_Element(driver : webdriver.Chrome, by, value : str) -> WebElement:
    while True:
        try:
            element = driver.find_element(by, value)
            break
        except:
            pass
        sleep(0.1)
    return element

def Find_Elements(driver : webdriver.Chrome, by, value : str) -> list[WebElement]:
    while True:
        try:
            elements = driver.find_elements(by, value)
            if len(elements) > 0:
                break
        except:
            pass
        sleep(0.1)
    return elements

def Send_Keys(element : WebElement, content : str):
    element.clear()
    for i in content:
        element.send_keys(i)
        sleep(0.1)

service = Service(executable_path="C:\chromedriver-win64\chromedriver.exe")   
options = Options()
options.add_experimental_option("debuggerAddress", "127.0.0.1:9001")
driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=options)

wb = Workbook()
sheet = wb.active
item = ['Product Link', 'Title', 'ebayID', 'ebay #', 'Price', 'Condition', 'Qty', 'Screen Size', 'Color', 'MPN', 'Aspect Ratio', 'Screen Finish', 'Maximum Resolution', 'Display Technology', 'Type', 'Brand', 'Compatible Brand', 'Compatible Model', 'Compatible Product Line', 'UPC']
for i in range(1, 21):
    sheet.cell(row = 1, column = i).value = item[i-1]

start_row = 2
item_num = 0
# the last number = last page + 1
for id in range(1, 8):
    driver.get(f'https://www.ebay.com/sch/i.html?_dkr=1&iconV2Request=true&_blrs=recall_filtering&_ssn=goodlucky2015&store_cat=0&store_name=goodlucky2015&_oac=1&_pgn={id}&rt=nc')
    product_items = driver.find_element(By.CLASS_NAME, "srp-results").find_elements(By.CLASS_NAME, "s-item__pl-on-bottom")
    output = []
    print('#' * 20, f'page {id}', '#' * 20)
    # print(product_items)
    for product_item in product_items:
        product_url = product_item.find_element(By.TAG_NAME, 'a').get_attribute('href')
        # print(product_url)
        output.append({"product link" : product_url})
        
    for item_index, item in enumerate(output):
        driver.get(item["product link"])
        sleep(0.1)
        try:
            title = driver.find_element(By.CLASS_NAME, 'x-item-title__mainTitle').text
            title_lower = title.lower()
            if 'motherboard' in title_lower or 'circuit board' in title_lower or 'battery' in title_lower or 'keyboard' in title_lower or 'desktop' in title_lower or 'antenna' in title_lower or 'wifi' in title_lower or 'front cover' in title_lower or 'back cover' in title_lower or 'bottom base' in title_lower or 'heatsink' in title_lower or 'ac adapter' in title_lower or 'ribbon' in title_lower or 'wire' in title_lower or 'led board' in title_lower or 'cable' in title_lower or 'base cover' in title_lower or 'converter' in title_lower or 'board led' in title_lower:
                continue
            elif 'lcd' in title_lower or 'led' in title_lower or 'panel' in title_lower or 'monitor' in title_lower or 'display' in title_lower:
                #link and title
                item_num += 1
                sheet.cell(row = start_row, column = 1).value = item["product link"]
                print('-' * 20, item_num, '-' * 20)
                print(f'Product URL : {item["product link"]}')
                sheet.cell(row = start_row, column = 2).value = title
                print(f'Title : {title}')
                
                # price
                try:
                    price = driver.find_element(By.CLASS_NAME, 'x-price-primary').text
                    if "US" not in price:
                        price = driver.find_element(By.CLASS_NAME, 'x-price-approx__price').text
                    sheet.cell(row = start_row, column = 3).value = price
                    print(f'Price : {price}')
                except:
                    pass
                
                # # condition
                # try:
                #     condition = 'Condition : ' + driver.find_element(By.CLASS_NAME, 'x-item-condition-text').find_element(By.CLASS_NAME, 'ux-textspans').text.strip()
                #     sheet.cell(row = start_row, column = 6).value = condition
                #     print(condition)
                # except:
                #     pass
                
                # # sold qty
                # try:
                #     sold = driver.find_element(By.CLASS_NAME, 'd-quantity__availability').text.split('/')
                #     sheet.cell(row = start_row, column = 7).value = sold[1]
                #     print(f'Sold : {sold[1]}')
                # except:
                #     pass
                
                # ebay #
                # try:
                #     ebayID = driver.find_element(By.CLASS_NAME, 'ux-layout-section--itemId').find_element(By.CLASS_NAME, 'ux-layout-section__textual-display--itemId').text
                #     sheet.cell(row = start_row, column = 4).value = ebayID
                #     print(f'ebay # : {ebayID}')
                # except:
                #     pass
                
                # specification
                try:
                    specifications = driver.find_elements(By.CLASS_NAME, 'ux-layout-section-evo__col')
                    for specification in specifications:
                        # if "Screen Size" in specification.text:
                        #     screen_size = 'Screen Size : ' + specification.text.replace('Screen Size', '').strip()
                        #     sheet.cell(row = start_row, column = 8).value = screen_size
                        #     print(screen_size)
                            
                        # elif "Color" in specification.text:
                        #     color = 'Color : ' + specification.text.replace('Color', '').strip()
                        #     sheet.cell(row = start_row, column = 9).value = color
                        #     print(color)
                        if "MPN" in specification.text:
                            mpn = 'MPN : ' + specification.text.replace('MPN', '').strip()
                            sheet.cell(row = start_row, column = 4).value = mpn
                            print(mpn)
                        
                        # elif "Aspect Ratio" in specification.text:
                        #     aspect_ratio = 'Aspect Ratio : ' + specification.text.replace('Aspect Ratio', '').strip()
                        #     sheet.cell(row = start_row, column = 11).value = aspect_ratio
                        #     print(aspect_ratio)
                        
                        # elif "Screen Finish" in specification.text:
                        #     screen_finish = 'Screen Finish : ' + specification.text.replace('Screen Finish', '').strip()
                        #     sheet.cell(row = start_row, column = 12).value = screen_finish
                        #     print(screen_finish)
                        
                        # elif "Maximum Resolution" in specification.text:
                        #     max_resolution = 'Maximum Resolution : ' + specification.text.replace('Maximum Resolution', '').strip()
                        #     sheet.cell(row = start_row, column = 13).value = max_resolution
                        #     print(max_resolution)
                        
                        # elif "Display Technology" in specification.text:
                        #     display_tech = 'Display Technology : ' + specification.text.replace('Display Technology', '').strip()
                        #     sheet.cell(row = start_row, column = 14).value = display_tech
                        #     print(display_tech)
                        
                        # elif "Type" in specification.text:
                        #     type = 'Type : ' + specification.text.replace('Type', '').strip()
                        #     sheet.cell(row = start_row, column = 15).value = type
                        #     print(type)
                        
                        # elif "Compatible Brand" in specification.text:
                        #     compatible_brand = 'Compatible Brand : ' + specification.text.replace('Compatible Brand', '').strip()
                        #     sheet.cell(row = start_row, column = 16).value = compatible_brand
                        #     print(compatible_brand)
                        
                        # elif "Brand" in specification.text:
                        #     brand = 'Brand : ' + specification.text.replace('Brand', '').strip()
                        #     sheet.cell(row = start_row, column = 17).value = brand
                        #     print(brand)
                        
                        elif "Compatible Models" in specification.text:
                            compatible_model = 'Compatible Model : ' + specification.text.replace('Compatible Models', '').strip()
                            sheet.cell(row = start_row, column = 5).value = compatible_model
                            print(compatible_model)
                        
                        elif "Compatible Model" in specification.text:
                            compatible_model = 'Compatible Model : ' + specification.text.replace('Compatible Model', '').strip()
                            sheet.cell(row = start_row, column = 5).value = compatible_model
                            print(compatible_model)
                        
                        # elif "Compatible Product Line" in specification.text:
                        #     compatible_product_line = 'Compatible Product Line : ' + specification.text.replace('Compatible Product Line', '').strip()
                        #     sheet.cell(row = start_row, column = 19).value = compatible_product_line
                        #     print(compatible_product_line)
                        
                        # elif "UPC" in specification.text:
                        #     upc = 'UPC : ' + specification.text.replace('UPC', '').strip()
                        #     sheet.cell(row = start_row, column = 20).value = upc
                        #     print(upc)
                except:
                    pass
                
                wb.save('output_goodlucky2015.xlsx')
                start_row += 1
                
                with open('item_url.json', 'w') as file:
                    json.dump(output, file)
        except:
            pass